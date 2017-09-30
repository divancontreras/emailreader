import win32com.client
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
import datetime
import time
import copy
from openpyxl.styles import Alignment

#book = openpyxl.load_workbook('\\\\WSMX02402FP\\Shared\\IMP-EXP\INTERNATIONAL TRADE\\RK & Activos\\EXPORT P2\\2017\HORARIOS EXP P2.xlsx')
book = openpyxl.load_workbook('HORARIOS EXP P2.xlsx')

outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
ws = book.worksheets[6]

inbox = outlook.GetDefaultFolder(6) # "6" refers to the index of a folder - in this case,
                                    # the inbox. You can change that number to reference                              # any other folder
messages = inbox.Items
print(inbox.Name)
##GET THE NEXT ROW COLUMN
for row in range(900,ws.max_row):
	if ws.cell(row=row,column=1).value == None:
		break

## Define what the filters will be
## This are variables that are compared on the if condition

emailGet = "expeditors"
subjectGet = 'SCHNEIDER ELECTRIC P2'
subject2Get = 'RE:'
bodyGet = 'attached entry'
 		##Iteration through all the messages on the default folder INBOX
 		##Condition that will filter the desire emails

for message in messages:
	if message.Class == 43:
	    if message.SenderEmailType == 'EX':
	    	msgSender = message.Sender.GetExchangeUser().PrimarySmtpAddress
	    else:
	   		msgSender = message.SenderEmailAddress
	if subjectGet in message.Subject and emailGet in msgSender and (subject2Get in message.Subject):
		rRow = [False] * 10000
		for x in range(1,row):
			if str(ws.cell(row=x,column=8).value) in message.Subject and rRow[x] != True:
				rRow[x] = True
				senton = str(message.SentOn)
				print(senton)
				ws.cell(row=x,column=18).value = senton.split()[0]
				ws.cell(row=x,column=19).value = senton.split()[1]

				print("______AGREGADO")

book.save('HORARIOS EXP P2.xlsx')
#book.save('\\\\WSMX02402FP\\Shared\\IMP-EXP\INTERNATIONAL TRADE\\RK & Activos\\EXPORT P2\\2017\HORARIOS EXP P2.xlsx')
print("DOCUMENTO GUARDADO")