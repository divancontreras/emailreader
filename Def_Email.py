import win32com.client
import openpyxl
from openpyxl.styles import Font
import datetime
import time
from openpyxl.styles import Alignment
from datetime import timedelta
from ClassEmailBox import *
from ClassExcelBox import *


excelfile = '\\\\WSMX02402FP\\Shared\\IMP-EXP\INTERNATIONAL TRADE\\RK & Activos\\EXPORT P2\\2017\HORARIOS EXP P2.xlsx'
#excelfile = 'HORARIOS EXP P2.xlsx'
excelbackup = '\\\\WSMX02402FP\\Shared\\IMP-EXP\INTERNATIONAL TRADE\\RK & Activos\\EXPORT P2\\2017\HORARIOS EXP P2 - BACKUP.xlsx'
dateformat = '%m/%d/%Y'

def SaveDoc(book):
    status = True
    while status:
        try:
            book.save(excelfile)
            status = False
        except:
            print("ERROR, EXCEL FILE OPEN, TRYING AGAIN IN 10 SECS!")
            status = True
            time.sleep(10)
    print("DOCUMENTO GUARDADO")


def InitEnv():
    print("Buscando cajas nuevas para agregar..." + "\n")
    book = openpyxl.load_workbook(excelfile)
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace(
        "MAPI")
    ws = book.worksheets[5]
    messages = (outlook.GetDefaultFolder(6).Items)  # "6" refers to the index of a folder "Inbox"
    icont = 0
    messages = list(reversed(messages))[0:300]
    cajas_folder = outlook.GetDefaultFolder(6).Folders.Item("Sistema de Cajas")
    IterInbox(ws, messages, cajas_folder, book)


def getRow(ws):
    for row in range(1000, ws.max_row):
        if ws.cell(row=row, column=1).value == None:
            return row


def CheckRepeat(ws, row, Sello, Caja, message):
    # This function will check if the box already exists.
    for xrow in range(row, row-30, -1):
        localsello = str(ws.cell(row=xrow, column=7).value)
        localbox = ws.cell(row=xrow, column=8).value
        ## VERIFICA IGUALDAD EN CAMPOS, Y QUE NO HAYA PALABRAS ADICIONALES EN EL ASUNTO.
        ## SI HAY PALABRAS ADICIONALES SIGNIFICA MODIFICACION Y NO SE REPITE.
        if Sello == localsello and Caja == localbox and len(str(message.subject).split()) == 7:
            return True

    return False


def FormatCells(ws, row):
    # This functions is only used to keep the same format in the document.
    for x in range(1, 21):
        cell = ws.cell(row=row, column=x)
        cell.font = Font(name="Calibri", size=9)
        cell.alignment = Alignment(horizontal='center', vertical='bottom')
        if (x == 3 or x == 4 or x == 14):
            cell.font = Font(name="Calibri", size=9, bold=True)


def IterInbox(ws, messages, cajas_folder, book):
    email = "AdminCajasMty"
    subject = 'P2- Reporte de Caja'
    subjectAd = 'NACIONAL'
    row = getRow(ws)
    rowAux = row
    for message in messages:
        if subject in message.Subject and email in message.SenderEmailAddress and not (
                subjectAd in message.Subject):
            if getData(message, ws, row):
                row += 1
            message.Move(cajas_folder)
    if rowAux != row:
        SaveDoc(book)
    else:
        print("NO BOXES WERE FOUND!")


def InserData(Caja, Sello, TipoCaja, Destino, Salida, ws, fecha, hora, row, message):
    ##    Fill the excel cells with the info from the email.
    ##    Each cell will be filled invidivually with the especific value
    if CheckRepeat(ws, row, Sello, Caja, message):
        return False
    else:
        fecha = fecha.split('-')
        formdate = fecha[1] + "/" + fecha[2] + "/" + fecha[0]
        today = datetime.date.today()
        ws.cell(
            row=row, column=1).value = datetime.datetime.now().isocalendar()[1]
        ws.cell(row=row, column=2).value = (today.strftime('%b')).upper()
        ws.cell(row=row, column=3).value = today.strftime("%#m/%#d/%Y")

        ##    This wil manage the Remesa number by the week of the year.
        if ((ws.cell(row=row - 1, column=1).value) !=
            (datetime.datetime.now().isocalendar()[1])):
            ws.cell(row=row, column=4).value = 1
        else:
            for xrow in range(row - 1, 0, -1):
                if ws.cell(row=xrow, column=4).value != None:
                    ws.cell(
                        row=row, column=4).value = ws.cell(
                            row=xrow, column=4).value + 1
                    break

        ##    Fill the excell cells with the info from the email.
        ##    Each cell will be filled invidivually with the especific value

        ws.cell(row=row, column=7).value = Sello
        ws.cell(row=row, column=8).value = Caja

        ## Specify when they are Dedicated boxes

        if 'HG' in TipoCaja:
            ws.cell(row=row, column=9).value = "DEDICADO"
        else:
            ws.cell(row=row, column=9).value = TipoCaja

        ws.cell(row=row, column=10).value = Destino.upper()
        t = time.strptime(Salida, "%H:%M")
        ws.cell(row=row, column=11).value = time.strftime("%I:%M %p", t)
        ws.cell(row=row, column=14).value = formdate
        ws.cell(row=row, column=15).value = hora[:8]
        FormatCells(ws, row)

        return True


def getData(message, ws, row):
    #Read the email and parse the key data
    #This variables are keeping the email data
    body = str(message.body).split()
    Caja = body[1]
    Sello = body[4]
    TipoCaja = message.Subject.split()[4]
    fecha = str(message.SentOn).split()[0]
    hora = str(message.SentOn).split()[1]
    # Some exceptions
    if ('Chino' in message.body):
        Destino = body[10] + " " + body[11]
        Salida = body[13]
    else:
        Destino = body[10]
        Salida = body[12]
    ##InserData() return a bool that states if the Data was successfully inserted
    if (InserData(Caja, Sello, TipoCaja, Destino, Salida, ws, fecha, hora,row,message)):
        print("_______BOX SAVED_______" + "\n")
        print(Caja + " " + Sello + " " + Destino + " " + Salida + " " +
              TipoCaja)
        print("___________________________" + "\n")
        return True
    ##If tha data was repeated the following will be displayed
    else:
        print("_______BOX REPEATED________" + "\n")
        print(Caja + " " + Sello + " " + Destino + " " + Salida + " " + TipoCaja)
        print("____________________________" + "\n")
        return False


def autoUpdateBox(timer):
    cont = 0
    timing = timer/60
    while True:
        book = openpyxl.load_workbook(excelfile)
        cont += 1
        InitEnv()
        if (cont == 30):
            status = True
            while status:
                try:
                    book.save(excelbackup)
                    status = False
                except:
                    print(
                        "ERROR, BACKUP OPEN!, CLOSE TO SAVE, TRYING EVERY 30SECS."
                    )
                    status = True
                    time.sleep(30)
            print("Documento Bakcup actualizado" + "\n")
            cont = 0
        
        print(f"Actualizacion cada {timing} minutos..." + "\n")
        time.sleep(timer)


def InitConf(EmailData):
    book = openpyxl.load_workbook(excelfile)
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    ws = book.worksheets[5]
    if 'schneider' in EmailData.getEmailKey():
        messages = outlook.GetDefaultFolder(5).Items
    else:
        messages = outlook.GetDefaultFolder(6).Items
    messages = list(reversed(messages))[0:300]
    IterConf(ws, messages, book, EmailData)


def IterConf(ws, messages, book, EmailData):
    row = getRow(ws)
    for message in messages:
        if message.Class == 43 :
            if message.SenderEmailType == "EX" :
                if 'XF' in str(message.Subject) and EmailData.getEmailKey() in message.Sender.GetExchangeUser().PrimarySmtpAddress:
                    getEmailData(ws, message, row, EmailData)
            else:
                if 'XF' in str(message.Subject) and EmailData.getEmailKey() in message.SenderEmailAddress:
                    getEmailData(ws, message, row, EmailData)
    if not(EmailData.getIteration()):
        SaveDoc(book)

def getEmailData(ws, message, row, EmailData):
    subject = str(message.Subject).upper()
    body = str(message.body).upper()
    if searchInvoice(subject, EmailData) :
        EmailData.setFecha(parseOldFormat(str(message.SentOn)))
        EmailData.setRow(getInvoiceRow(row, ws, EmailData.getFactura()))
        setBoxConfirmation(ws, message, EmailData)


def setBoxConfirmation(ws, message, EmailData):
    if EmailData.getRow() != None:
        Date = EmailData.getFecha().strftime("%m/%d/%Y")  
        Hour = EmailData.getFecha().strftime("%H:%M:%S")
        ws.cell(row = EmailData.getRow(), column = EmailData.getDateColumn()).value = Date
        ws.cell(row = EmailData.getRow(), column = EmailData.getHourColumn()).value = Hour


def searchInvoice(message, EmailData):
    for x in message.split():
        if 'XF' in x:
            EmailData.setFactura(x)
            return True
    return False


def getInvoiceRow(row, ws, factura):
    for x in range(row,1,-1):
        if ws.cell(row=x, column=6).value == factura :
            return x
    return None

def parseOldFormat(fecha):
    if fecha != 'None' :
        fecha = str(fecha)
        if "+" in fecha:
            fecha = fecha.split('+')[0]
        if '/' in fecha:
            return fecha
        if '-' in fecha and not(':' in fecha):
            return datetime.datetime.strptime(fecha,"%Y-%m-%d")            
        return datetime.datetime.strptime(fecha,"%Y-%m-%d %H:%M:%S")

def test():
    with open('config.txt', 'r') as content_file:
        icont = 0
        content = content_file.read().split('"')
        for x in content:
            if not ('#' in x):
                print(str(icont) + " - " + x)
            icont += 1
    
def main():
    while True:
        with open('config.txt', 'r') as content_file:
            content = content_file.read().split('"')
            print(f"""
            1. Actualizar Cajas Excel
            2. Actualizado automatico 
            3. Actualizar Confirmacion {content[9]}
            4. Actualizar Confirmacion {content[17]}
            5. Actualizar Confirmacion {content[25]}
            6. Actualizar Confirmaciones Todas ({content[9]},{content[17]},{content[25]})
            7. Exit/Quit
            9. Selecciona si es tu primera vez ejecutando el programa.
            """)
            ans = input("Seleccione una numero: ")
            if ans == '1':

                BoxData = BoxExcel(content[1],content[3],content[5])
                InitEnv()

            if ans == '2':
                time = int(content[7])/60
                print("\n" + f"Actualizacion cada {time} minutos..." + "\n")
                autoUpdateBox(int(content[7]))

            if ans == '3':
                print("\n" + f"Buscando confirmaciones {content[9]}..." + "\n")
                EmailData = EmailBox(content[11],int(content[13]),int(content[15]))
                InitConf(EmailData)

            elif ans == '4':
                print("\n" + f"Buscando confirmaciones {content[17]}..." + "\n")            
                EmailData = EmailBox(content[19],int(content[21]),int(content[23]))
                InitConf(EmailData)

            elif ans == '5':
                print("\n" + f"Buscando confirmaciones {content[25]}..." + "\n")
                EmailData = EmailBox(content[27],int(content[29]),int(content[31]))
                InitConf(EmailData)

            elif ans == '6':
                print("\n" + f"Buscando confirmaciones {content[9]}..." + "\n")
                EmailData = EmailBox(content[11],int(content[13]),int(content[15]),True)
                InitConf(EmailData)
                print(f"Buscando confirmaciones {content[17]}..." + "\n")            
                EmailData = EmailBox(content[19],int(content[21]),int(content[23]),True)
                InitConf(EmailData)
                print(f"Buscando confirmaciones {content[25]}..." + "\n")
                EmailData = EmailBox(content[27],int(content[29]),int(content[31]),False)
                InitConf(EmailData)


            elif ans == '9':
                print("Adios")
                exit()
            elif ans == '7':
                test()
main()
