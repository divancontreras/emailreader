import win32com.client
import openpyxl
from openpyxl.styles import Font
import datetime
import time
from openpyxl.styles import Alignment
from datetime import timedelta
from ClassEmailBox import *
from ClassExcelBox import *


excelfile = '\\\\WSMX02402FP\\Shared\\IMP-EXP\INTERNATIONAL TRADE\\RK & Activos\\EXPORT P2\\2017\HORARIOS EXP P2.xlsx'
#excelfile = 'HORARIOS EXP P2.xlsx'
excelbackup = '\\\\WSMX02402FP\\Shared\\IMP-EXP\INTERNATIONAL TRADE\\RK & Activos\\EXPORT P2\\2017\HORARIOS EXP P2 - BACKUP.xlsx'
dateformat = '%m/%d/%Y'

def SaveDoc(book):
    status = True
    while status:
        try:
            book.save(excelfile)
            status = False
        except:
            print("ERROR, EXCEL FILE OPEN, TRYING AGAIN IN 10 SECS!")
            status = True
            time.sleep(10)
    print("DOCUMENTO GUARDADO")


def InitEnv(BoxData):
    print("Buscando cajas nuevas para agregar..." + "\n")
    book = openpyxl.load_workbook(excelfile)
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace(
        "MAPI")
    ws = book.worksheets[5]
    messages = (outlook.GetDefaultFolder(6).Items)  # "6" refers to the index of a folder "Inbox"
    messages = list(messages)
    messages = messages[len(messages)-350:]
    cajas_folder = outlook.GetDefaultFolder(6).Folders.Item("Sistema de Cajas")
    IterInbox(ws, messages, cajas_folder, book, BoxData)


def getRow(ws):
    for row in range(1000, ws.max_row):
        if ws.cell(row=row, column=1).value == None:
            return row


def CheckRepeat(ws, message, BoxData):
    # This function will check if the box already exists.
    for xrow in range(BoxData.getRow(), BoxData.getRow()-50, -1):
        localsello = str(ws.cell(row=xrow, column=7).value)
        localbox = ws.cell(row=xrow, column=8).value
        ## VERIFICA IGUALDAD EN CAMPOS.
        if BoxData.getSello() == localsello and BoxData.getCaja() == localbox:
            return True
    return False


def FormatCells(ws, row):
    # This functions is only used to keep the same format in the document.
    for x in range(1, 21):
        cell = ws.cell(row=row, column=x)
        cell.font = Font(name="Calibri", size=9)
        cell.alignment = Alignment(horizontal='center', vertical='bottom')
        if (x == 3 or x == 4 or x == 14):
            cell.font = Font(name="Calibri", size=9, bold=True)


def IterInbox(ws, messages, cajas_folder, book, BoxData):
    BoxData.setRow(getRow(ws))
    rowAux = BoxData.getRow()
    for message in messages:
        if BoxData.getSubject() in message.Subject and BoxData.getEmail() in message.SenderEmailAddress and not (
                BoxData.getAdSubject() in message.Subject):
            if getData(message, ws, BoxData):
                BoxData.setRow(BoxData.getRow()+1)
            message.Move(cajas_folder)
    if rowAux != BoxData.getRow():
        SaveDoc(book)
    else:
        print("NO BOXES WERE FOUND!")


def InserData(BoxData, ws, message):
    ##    Fill the excel cells with the info from the email.
    ##    Each cell will be filled invidivually with the especific value
    if CheckRepeat(ws, message, BoxData):
        return False
    else:
        BoxData.setFecha((parseOldFormat(BoxData.getFecha())).strftime(dateformat))
        today = datetime.date.today()
        ws.cell(row=BoxData.getRow(), column=1).value = datetime.datetime.now().isocalendar()[1]
        ws.cell(row=BoxData.getRow(), column=2).value = (today.strftime('%b')).upper()
        ws.cell(row=BoxData.getRow(), column=3).value = today.strftime("%#m/%#d/%Y")

        ##    This wil manage the Remesa number by the week of the year.
        if ((ws.cell(row=BoxData.getRow() - 1, column=1).value) !=
            (datetime.datetime.now().isocalendar()[1])):
            ws.cell(row=BoxData.getRow(), column=4).value = 1
        else:
            for xrow in range(BoxData.getRow() - 1, 0, -1):
                if ws.cell(row=xrow, column=4).value != None:
                    ws.cell(
                        row=BoxData.getRow(), column=4).value = ws.cell(
                            row=xrow, column=4).value + 1
                    break

        ##    Fill the excell cells with the info from the email.
        ##    Each cell will be filled invidivually with the especific value

        ws.cell(row=BoxData.getRow(), column=7).value = BoxData.getSello()
        ws.cell(row=BoxData.getRow(), column=8).value = BoxData.getCaja()

        ## Specify when they are Dedicated boxes

        if 'HG' in BoxData.getTipoCaja():
            ws.cell(row=BoxData.getRow(), column=9).value = "DEDICADO"
        else:
            ws.cell(row=BoxData.getRow(), column=9).value = BoxData.getTipoCaja()

        ws.cell(row=BoxData.getRow(), column=10).value = (BoxData.getDestino()).upper()
        t = time.strptime(BoxData.getSalida(), "%H:%M")
        ws.cell(row=BoxData.getRow(), column=11).value = time.strftime("%I:%M %p", t)
        ws.cell(row=BoxData.getRow(), column=14).value = BoxData.getFecha()
        ws.cell(row=BoxData.getRow(), column=15).value = BoxData.getHora()[:8]
        FormatCells(ws, BoxData.getRow())

        return True


def getData(message, ws, BoxData):
    #Read the email and parse the key data
    #This variables are keeping the email data
    BoxData.setBody(str(message.body).split())
    BoxData.setCaja(BoxData.getBody()[1])
    BoxData.setSello(BoxData.getBody()[4])
    BoxData.setTipoCaja(str(message.Subject).split()[4])
    BoxData.setFecha(str(message.SentOn).split()[0])
    BoxData.setHora(str(message.SentOn).split()[1])

    # Some exceptions
    if ('Chino' in BoxData.getBody()):
        BoxData.setDestino(BoxData.getBody()[10] + " " + BoxData.getBody()[11])
        BoxData.setSalida(BoxData.getBody()[13])
    else:
        BoxData.setDestino(BoxData.getBody()[10])
        BoxData.setSalida(BoxData.getBody()[12])
    ##InserData() return a bool that states if the Data was successfully inserted
    if (InserData(BoxData, ws, message)):
        print(f"""
        ______________BOX SAVED______________
        {BoxData.getCaja()} {BoxData.getSello()} {BoxData.getDestino()} {BoxData.getSalida()} {BoxData.getTipoCaja()}
        _____________________________________
        """)
        return True
    ##If tha data was repeated the following will be displayed
    else:
        print(f"""
        ______________BOX REPEATED______________
        {BoxData.getCaja()} {BoxData.getSello()} {BoxData.getDestino()} {BoxData.getSalida()} {BoxData.getTipoCaja()}
        _____________________________________
        """)


def autoUpdateBox(timer):
    cont = 0
    timing = timer/60
    while True:
        book = openpyxl.load_workbook(excelfile)
        cont += 1
        InitEnv()
        if (cont == 30):
            saveBackUp(book)
            cont = 0   
        print(f"Actualizacion cada {timing} minutos..." + "\n")
        time.sleep(timer)


def saveBackUp(book):
    status = True
    while status:
        try:
            book.save(excelbackup)
            status = False
        except:
            print(
                "ERROR, BACKUP OPEN!, CLOSE TO SAVE, TRYING EVERY 30SECS."
            )
            status = True
            time.sleep(30)
    print("Documento Bakcup actualizado" + "\n")


def InitConf(EmailData):
    book = openpyxl.load_workbook(excelfile)
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    ws = book.worksheets[5]
    if 'schneider' in EmailData.getEmailKey():
        messages = outlook.GetDefaultFolder(5).Items
    else:
        messages = outlook.GetDefaultFolder(6).Items
    messages = list(messages)
    messages = messages[len(messages)-600:]
    IterConf(ws, messages, book, EmailData)


def IterConf(ws, messages, book, EmailData):
    row = getRow(ws)
    for message in messages:
        if message.Class == 43 :
            if message.SenderEmailType == "EX" :
                if 'XF' in str(message.Subject) and EmailData.getEmailKey() in message.Sender.GetExchangeUser().PrimarySmtpAddress:
                    getEmailData(ws, message, row, EmailData)
            else:
                if 'XF' in str(message.Subject) and EmailData.getEmailKey() in message.SenderEmailAddress:
                    getEmailData(ws, message, row, EmailData)
    if not(EmailData.getIteration()):
        SaveDoc(book)

def getEmailData(ws, message, row, EmailData):
    subject = str(message.Subject).upper()
    body = str(message.body).upper()
    if searchInvoice(subject, EmailData) :
        EmailData.setFecha(parseOldFormat(str(message.SentOn)))
        EmailData.setRow(getInvoiceRow(row, ws, EmailData.getFactura()))
        setBoxConfirmation(ws, message, EmailData)


def setBoxConfirmation(ws, message, EmailData):
    if EmailData.getRow() != None:
        Date = EmailData.getFecha().strftime("%m/%d/%Y")  
        Hour = EmailData.getFecha().strftime("%H:%M:%S")
        ws.cell(row = EmailData.getRow(), column = EmailData.getDateColumn()).value = Date
        ws.cell(row = EmailData.getRow(), column = EmailData.getHourColumn()).value = Hour


def searchInvoice(message, EmailData):
    for x in message.split():
        if 'XF' in x:
            EmailData.setFactura(x)
            return True
    return False


def getInvoiceRow(row, ws, factura):
    for x in range(row,1,-1):
        if ws.cell(row=x, column=6).value == factura :
            return x
    return None

def parseOldFormat(fecha):
    if fecha != 'None' :
        fecha = str(fecha)
        if "+" in fecha:
            fecha = fecha.split('+')[0]
        if '/' in fecha:
            return fecha
        if '-' in fecha and not(':' in fecha):
            return datetime.datetime.strptime(fecha,"%Y-%m-%d")            
        return datetime.datetime.strptime(fecha,"%Y-%m-%d %H:%M:%S")

    
def main():
    print("""
    Email reader, a python script for parsing email data. 
    Author: Diego Contreras at Diego.conville@gmail.com 
    """)
    while True:
        with open('config.txt', 'r') as content_file:
            content = content_file.read().split('"')
            print(f"""
            1. Actualizar Cajas Excel
            2. Actualizado automatico 
            3. Actualizar Confirmacion {content[9]}
            4. Actualizar Confirmacion {content[17]}
            5. Actualizar Confirmacion {content[25]}
            6. Actualizar Confirmaciones Todas ({content[9]},{content[17]},{content[25]})
            7. Exit/Quit
            """)
            ans = input("Seleccione una numero: ")
            if ans == '1':
    
                BoxData = BoxExcel(content[1],content[3],content[5])
                InitEnv(BoxData)

            if ans == '2':
                time = int(content[7])/60
                print("\n" + f"Actualizacion cada {time} minutos..." + "\n")
                autoUpdateBox(int(content[7]))

            if ans == '3':
                print("\n" + f"Buscando confirmaciones {content[9]}..." + "\n")
                EmailData = EmailBox(content[11],int(content[13]),int(content[15]))
                InitConf(EmailData)

            elif ans == '4':
                print("\n" + f"Buscando confirmaciones {content[17]}..." + "\n")            
                EmailData = EmailBox(content[19],int(content[21]),int(content[23]))
                InitConf(EmailData)

            elif ans == '5':
                print("\n" + f"Buscando confirmaciones {content[25]}..." + "\n")
                EmailData = EmailBox(content[27],int(content[29]),int(content[31]))
                InitConf(EmailData)

            elif ans == '6':
                print("\n" + f"Buscando confirmaciones {content[9]}..." + "\n")
                EmailData = EmailBox(content[11],int(content[13]),int(content[15]),False)
                InitConf(EmailData)
                print(f"Buscando confirmaciones {content[17]}..." + "\n")            
                EmailData = EmailBox(content[19],int(content[21]),int(content[23]),False)
                InitConf(EmailData)
                print(f"Buscando confirmaciones {content[25]}..." + "\n")
                EmailData = EmailBox(content[27],int(content[29]),int(content[31]),False)
                InitConf(EmailData)

            elif ans == '7':
                exit()
main()
