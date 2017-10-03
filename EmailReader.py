import win32com.client
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
import datetime
import time
import copy
from openpyxl.styles import Alignment

while True:
	book = openpyxl.load_workbook('\\\\WSMX02402FP\\Shared\\IMP-EXP\INTERNATIONAL TRADE\\RK & Activos\\EXPORT P2\\2017\HORARIOS EXP P2.xlsx')
	outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
	ws = book.worksheets[6]

	inbox = outlook.GetDefaultFolder(6) # "6" refers to the index of a folder - in this case,
	                                    # the inbox. You can change that number to reference                              # any other folder
	messages = inbox.Items
	cajas_folder = inbox.Folders.Item("Sistema de Cajas")
	print(inbox.Name)
	print(cajas_folder.Name)
	today = datetime.date.today()
	##GET THE NEXT ROW COLUMN
	for row in range(900,ws.max_row):
		if ws.cell(row=row,column=1).value == None:
			break

	## Define what the filters will be
	## This are variables that are compared on the if condition

	email = "AdminCajasMty"
	subject = 'P2- Reporte de Caja'
	subjectAd = 'NACIONAL'

	 		##Iteration through all the messages on the default folder INBOX
	 		##Condition that will filter the desire emails

	for message in messages:
		print(message.Subject)
		if subject in message.Subject and email in message.SenderEmailAddress and not(subjectAd in message.Subject):


			#Read the Email and get the key data from it.
			#This are the variables that will have the data of the email

			body = message.body
			Caja = body.split()[1]
			Sello = body.split()[4]
			TipoCaja = message.Subject.split()[4]

			if('Chino' in message.body):
				Destino = body.split()[10] + " " + body.split()[11]
				Salida = body.split()[13]
			else:
				Destino = body.split()[10]
				Salida = body.split()[12]

			print(message.subject)
			print(Caja)
			print(Sello)
			print(Destino)
			print(Salida)
			print(TipoCaja)		
			##	Fill the excell cells with the info from the email.
			##	Each cell will be filled invidivually with the especific value
			ws.cell(row=row, column=1).value = datetime.datetime.now().isocalendar()[1]
			ws.cell(row=row, column=2).value = (today.strftime('%b')).upper()
			ws.cell(row=row, column=3).value = today.strftime("%#m/%#d/2017")

			##	This wil manage the Remesa number by the week of the year.
			if((ws.cell(row=row-1, column = 1).value) != (datetime.datetime.now().isocalendar()[1])):
				ws.cell(row=row, column=4).value = 1
			else:
				ws.cell(row=row, column=4).value = ws.cell(row=row-1, column=4).value + 1

			##	Fill the excell cells with the info from the email.
			##	Each cell will be filled invidivually with the especific value

			ws.cell(row=row, column=7).value = Sello
			ws.cell(row=row, column=8).value = Caja
			
			## Specify when they are Dedicated boxes

			if 'HG' in TipoCaja:
				ws.cell(row=row, column=9).value = "DEDICADO"
			else:
				ws.cell(row=row, column=9).value = TipoCaja

			ws.cell(row=row, column=10).value = Destino.upper()
			t = time.strptime( Salida , "%H:%M")
			ws.cell(row=row, column=11).value = time.strftime( "%I:%M %p", t)
			ws.cell(row=row, column=14).value = str(message.SentOn).split()[0]
			ws.cell(row=row, column=15).value = str(message.SentOn).split()[1]


			for x in range(1,21):
				cell = ws.cell(row=row, column=x) 
				cell.font = Font(name="Calibri", size = 9)
				cell.alignment = Alignment(horizontal='center', vertical='bottom')
				if(x == 3 or x == 4):
					cell.font = Font(name="Calibri", size = 9, bold = True)
			print("Caja Guardada")
			print("___________________________")
			message.Move(cajas_folder) 
			row=row+1
		# elif '// SCHNEIDER ELECTRIC P2' in message.Subject and 'expeditors' in message.SenderEmailAddress and 'Please see attached' in message.body:
		# 	caja = message.Subject.split()[2]
		# 	for x in range(row, row-15 , -1):
		# 		if(ws.cell(row=x, column=8).value == caja):
		# 			ws.cell(row=row, column=19).value == message.SentOn
		# 			print(caja + "Actualizada")
		# elif '// SCHNEIDER ELECTRIC P2' in message.Subject and 'dawvidal' in message.SenderEmailAddress and 'PEDIMENTO LISTO' in message.body:
		# 	caja = message.Subject.split()[2]
		# 	for x in range(row, row-15 , -1):
		# 		if(ws.cell(row=x, column=8).value == caja):
		# 			ws.cell(row=row, column=21).value == message.SentOn
		# 			print(caja + "Actualizada")\
		status = True
	while status:
		try:
			book.save('\\\\WSMX02402FP\\Shared\\IMP-EXP\INTERNATIONAL TRADE\\RK & Activos\\EXPORT P2\\2017\HORARIOS EXP P2.xlsx')
			status = False
		except:
			print("ERROR, ARCHIVO DE HORARIOS ABIERTO!, CIERRE PARA GUARDAR, VOLVIENDO A INTENTAR EN 30 SEGUNDOS.")
			status = True
			time.sleep(30)
	print("DOCUMENTO GUARDADO")
	time.sleep(400)
	#Input = input("Enter para refrescar / S para Salir / Se refresca automaticamente cada 5 minutos: ")
	#if end - start > expires_in:
	#	print("Actualizando...")
	#if(Input.lower() == 's'):
	#	exit()
